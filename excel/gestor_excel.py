"""
excel/gestor_excel.py
Gestiona comprobantes.xlsx
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

HEADERS    = ["Fecha", "Banco", "Titular receptor", "COD CTA",
              "Cliente", "Monto", "Titular de cuenta", "N° Operación", "Estado"]
COL_WIDTHS = [18, 16, 28, 10, 25, 14, 28, 22, 12]

FILL_HEADER   = PatternFill(start_color="1a7a4a", end_color="1a7a4a", fill_type="solid")
FILL_OK       = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
FILL_REVISION = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
FONT_HEADER   = Font(color="FFFFFF", bold=True)
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center")


class GestorExcel:
    def __init__(self, config: dict, log_callback):
        self.log = log_callback
        self.ruta = config["archivo_excel"]
        self.operaciones_existentes: set = set()

        if not os.path.exists(self.ruta):
            self._crear_excel()

    def _crear_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Comprobantes"
        ws.append(HEADERS)

        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER

        for i, w in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(self.ruta)
        self.log(f"📄 Excel creado: {self.ruta}")

    def cargar_operaciones_existentes(self):
        if not os.path.exists(self.ruta):
            self._crear_excel()
            return

        wb = load_workbook(self.ruta, read_only=True, data_only=True)
        ws = wb.active
        self.operaciones_existentes.clear()

        for row in ws.iter_rows(min_row=2, values_only=True):
            nro_op = row[7]
            if nro_op and str(nro_op).strip() not in ("XX", ""):
                self.operaciones_existentes.add(str(nro_op).strip())

        wb.close()
        self.log(f"ℹ {len(self.operaciones_existentes)} operaciones cargadas del Excel.")

    def es_duplicado(self, nro_operacion: str) -> bool:
        if not nro_operacion or nro_operacion.strip() in ("XX", ""):
            return False
        return nro_operacion.strip() in self.operaciones_existentes

    def registrar(self, registro: dict):
        estado = registro.get("estado", "Revisión")
        fill_fila = FILL_OK if estado == "OK" else FILL_REVISION

        fila = [
            registro.get("fecha", ""),
            registro.get("banco", ""),
            registro.get("titular_receptor", ""),
            registro.get("cod_cta", ""),
            registro.get("cliente", ""),
            registro.get("monto", ""),
            registro.get("titular", ""),
            registro.get("nro_operacion", ""),
            estado,
        ]

        # Manejo de archivo abierto
        ruta_guardar = self.ruta
        try:
            wb = load_workbook(self.ruta)
        except PermissionError:
            base, ext = os.path.splitext(self.ruta)
            ruta_guardar = base + "_temp" + ext
            self.log(f"⚠ Excel en uso → guardando en {ruta_guardar}")
            wb = load_workbook(self.ruta)

        ws = wb.active
        ws.append(fila)
        fila_idx = ws.max_row

        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=fila_idx, column=col).fill = fill_fila

        wb.save(ruta_guardar)

        nro_op = registro.get("nro_operacion", "")
        if nro_op and nro_op.strip() not in ("XX", ""):
            self.operaciones_existentes.add(nro_op.strip())

        self.log(f"💾 Guardado en Excel: fila {fila_idx}")